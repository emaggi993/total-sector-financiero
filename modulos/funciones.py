import pandas as pd
from modulos.fecha import Fecha
from openpyxl import load_workbook

def leer_archivo(file_name: str, sheet_name: str) -> pd.DataFrame:
    if not file_name or not sheet_name:
        raise Exception("Se debe enviar el nombre de archivo y nombre de hoja")
    try:
        archivo = pd.read_excel(file_name, sheet_name=sheet_name)
    except Exception as e:
        raise Exception("se ha encontrado un error "+str(e))
    else:
        return archivo
def obtener_datos(archivo: str, fecha: Fecha) -> dict[pd.DataFrame]:
    """
    Params:
        archivo: path del archivo de resumen de bancos o financieras
        fecha: fecha,

    Returns:
        Diccionario con los Dataframe con los datos de todas las hojas
    """
    try:
        workbook = load_workbook(archivo)
        sheet_names = workbook.sheetnames
        sheet_names.remove('historico_morosidad')
        dictionary = {}
        for sheet_name in sheet_names:
            # print(sheet_name)
            datos = pd.read_excel(archivo, sheet_name)
            datos.rename(columns={'anho':'anio'},
               inplace=True)
            datos = datos[(datos['mes']== int(fecha.mes)) & (datos['anio'] == int(fecha.anio))]
            dictionary[sheet_name]= datos
        # print(datos.keys())
        return dictionary
    
    except Exception as e:
        raise "Error en obtener datos ->"+ str(e)
    
