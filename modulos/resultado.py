from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from dateutil.relativedelta import relativedelta
from modulos.fecha import Fecha
from datetime import datetime
import pandas as pd
import traceback
title_format = {"bgColor" : "339966", "color": "FFFFFF"}
class Resultado():
    def __init__(self, archivo) -> None:
        self.path = archivo
        self.workbook = load_workbook(archivo)
    def set_bancos(self, cant: int, col: str, row: int = 4):
        self.sheet[col + str(row)].value = cant
    def pos_fecha( self,  fecha: Fecha) -> int:
        
        '''
        fecha: fecha buscada en el informe

        Return 
            posicion donde se encuentra la fecha o -1 si no esta
        '''
        try:
            df = pd.read_excel(self.path , sheet_name= self.sheet_name, header=None)
            self.df = df.copy()
            index = None
            if self.sheet_name== "BANCOS":
                index = df[df[0] == 'BANCOS'].index.values[0]
            elif self.sheet_name == "FINANCIERAS":
                index = df[df[0] == 'FINANCIERAS'].index.values[0]
            else:
                raise Exception("solo se acepta nombre de hoja BANCOS o FINANCIERAS <- resultado.Resultado.verificar_fecha")
            
            fechas = [ item.strftime("%d/%m/%Y") for item in df.iloc[index][1:].to_list()]
            if str(fecha) in fechas:
                return fechas.index(str(fecha)) + 1
            return -1
            
        except:
            raise traceback.print_exc()
    def cell_value(self, celda, value)-> None:
        self.sheet[celda].value = value
    def set_sheet(self, sheet_name):
        try:
            self.sheet = self.workbook[sheet_name]
            self.sheet_name = sheet_name
            self.data = pd.read_excel(self.path, sheet_name= sheet_name, header= None)
            
            self.max_row = self.sheet.max_row
            self.max_column = self.sheet.max_column
            self.letter_max_column = get_column_letter(self.max_column)
        except:
            self.sheet = None
            raise Exception("No existe la hoja")

    def get_sheet(self, sheet_name):

        return self.workbook[sheet_name]

    def set_titulo(self, title: str = None, title_column: str = None, row: int = 6)-> None:
        if title_column is None:
            title_column = get_column_letter( self.max_column + 1)
        print("set titulo ->", title_column)
        if title == None:
            title = self.sheet[get_column_letter(self.max_column)+str(row)].value + relativedelta(months=1 )
        if isinstance(title, Fecha):
            f = datetime( int(title.anio), int(title.mes), int(title.dia) )
        self.sheet[title_column + str(row)].value = f
        self.sheet[title_column + str(row)].number_format = "mmm-yy"
        self.sheet[title_column + str(row)].font = Font(color= title_format['color'], size=8, bold=True )
        self.sheet[title_column + str(row)].fill = PatternFill(start_color= title_format["bgColor"],end_color=title_format["bgColor"],fill_type = "solid")
        #339966

    def save(self):
        self.workbook.save(self.path)
    
    def get_fechas(self, row: int = 5):
        return self.data.iloc[row][1:].to_list()