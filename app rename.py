from modulos.fecha import Fecha
from modulos.funciones import leer_archivo
from modulos.bancos import Bancos
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta
from modulos.resultado import Resultado
from modulos.informe import datos_informe, bank_count
from openpyxl import load_workbook
import traceback

if __name__ == "__main__":
    # datos = leer_archivo(r"datos/BOLB_BANCOS.xlsx", "balances")
    fecha = Fecha("01/01/2022")
    datos = Bancos(r"datos/BOLB_BANCOS.xlsx", fecha)
    # datos = datos.filtro_fecha_cuenta(Fecha("01/08/2022"), "CAJA Y BANCOS")
    # print(datos)
    # wb = load_workbook(r'datos\Total Sistema Financiero Base.xlsx')
    r  = Resultado(r'datos\Total Sistema Financiero Base.xlsx')
    r.set_sheet('BANCOS')
    informe_total_libro = load_workbook(r'datos\Total Sistema Financiero Base.xlsx')
    hoja = informe_total_libro['BANCOS']
    pos_fecha = r.pos_fecha(fecha)
    print(pos_fecha)
    url_master= "Lista de cuentas.xlsx"
    url_datos= "./datos/BOLB_BANCOS.xlsx"
    result = datos_informe(url_master, url_datos, fecha)
    tipocambio = 6957
    f,c = result.shape
    if pos_fecha != -1:
        column = get_column_letter(pos_fecha + 1)
        colanterior = get_column_letter(pos_fecha )
    else:
        column = get_column_letter(r.max_column + 1)
        colanterior = get_column_letter(r.max_column )
    bancos = bank_count(url_datos, "balances", fecha)
    print(bancos, "<------------ cantidad de bancos en la fecha -> ", str(fecha))
    r.set_titulo(fecha, column)
    r.set_bancos(bancos, column)
    for fila in range(f):
        try:
            d = result.iloc[fila].dropna()
            # f = d[0]
            print("--"*30)
            print(column, d['fila'])
            if d['hoja'] == 'formula':
                value = "="+d[d['hoja']].replace("{col}", column).replace("{colanterior}", colanterior)
                print("="+d[d['hoja']].replace("{col}", column).replace("{colanterior}", colanterior))
            elif d['hoja']=='tipo-cambio':
                value = tipocambio
            else:
                value = d[d['hoja']]
                print(d[d['hoja']])
            r.cell_value(column+str( d['fila']) , value)
        except KeyError:
            # raise Exception("keyError")
            pass
        except:
            raise traceback.print_exc()


    r.save()


